import openpyxl
import json
from pathlib import Path

# Load Excel file
wb = openpyxl.load_workbook(r'../planilla de costos Mer.xlsx', data_only=True)

# Convert ingredients from LISTA DE PRECIOS sheet
def convert_ingredients():
    ws = wb['LISTA DE PRECIOS MAYORISTAS ']
    ingredients = []
    
    current_category = None
    for row in ws.iter_rows(min_row=6, values_only=True):  # Start at row 6 to skip headers
        # Skip empty rows
        if not any(row):
            continue
            
        # Check if this is a category header (has text in col 2, no number in col 3)
        if row[2] and (not row[3] or isinstance(row[3], str)):
            category_text = str(row[2]).strip().upper()
            # Skip headers like "CANTIDAD", "U.M.", etc
            if category_text in ['CANTIDAD', 'U.M.', 'PRECIO DE COMPRA', 'PRECIO UNITARIO', 'SECOS']:
                continue
            current_category = category_text
            continue
        
        # Skip if no name or name is a header
        if not row[2] or str(row[2]).strip() in ['CANTIDAD', 'Materia Prima']:
            continue
            
        name = str(row[2]).strip()
        quantity = row[3]
        unit = row[4]
        purchase_price = row[5]
        unit_price = row[7]
        
        # Try to convert to numbers
        try:
            qty = float(quantity) if quantity and not isinstance(quantity, str) else None
            pp = float(purchase_price) if purchase_price and not isinstance(purchase_price, str) else None
            up = float(unit_price) if unit_price and not isinstance(unit_price, str) else None
            
            if name and qty and unit:
                ingredients.append({
                    'id': f'ing_{len(ingredients) + 1}',
                    'name': name,
                    'category': current_category or 'Otros',
                    'quantity': qty,
                    'unit': str(unit).strip() if unit else 'unidades',
                    'purchasePrice': pp if pp else 0,
                    'unitPrice': up if up else (pp / qty if pp and qty else 0)
                })
        except (ValueError, TypeError):
            continue
    
    return ingredients

# Convert correction factors
def convert_correction_factors():
    ws = wb['FACTOR DE CORRECCION']
    factors = []
    
    # Start from row 3 (after headers)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        
        # Extract from each category column
        categories = [
            ('vegetables', 1, 2, 3),
            ('fruits', 4, 5, 6),
            ('meat', 7, 8, 9),
            ('other', 10, 11, 12)
        ]
        
        for cat_name, name_col, fc_col, waste_col in categories:
            name = row[name_col]
            fc = row[fc_col]
            waste = row[waste_col]
            
            if name and name != 'FC' and name != 'MERMA':  # Skip header rows
                try:
                    fc_value = float(fc) if fc and not isinstance(fc, str) else 1.0
                    waste_value = 0
                    if waste and waste != '–' and not isinstance(waste, str):
                        waste_value = float(waste) * 100
                    
                    factors.append({
                        'name': str(name).strip(),
                        'category': cat_name,
                        'correctionFactor': fc_value,
                        'wastePercentage': waste_value
                    })
                except (ValueError, TypeError):
                    continue
    
    return factors

# Convert nutritional info
def convert_nutritional_info():
    ws = wb['TABLA QUIMICA DE ALIMENTOS ']
    nutritional_data = []
    
    current_category = None
    current_subcategory = None
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        
        name = row[1]
        if not name:
            continue
        
        name_str = str(name).strip()
        
        # Check if category
        if name_str.startswith('1.') or name_str.startswith('2.') or name_str.startswith('3.'):
            current_category = name_str
            continue
        
        # Check if subcategory (contains "kcal")  
        if row[2] and str(row[2]).strip() == 'kcal / 100g':
            current_subcategory = name_str
            continue
        
        # Regular food item
        calories = row[2]
        carbs = row[3]
        protein = row[4]
        fat = row[5]
        
        if calories and isinstance(calories, (int, float)):
            nutritional_data.append({
                'name': name_str,
                'category': current_category or 'Sin categoría',
                'subcategory': current_subcategory or '',
                'calories': float(calories) if calories else 0,
                'carbs': float(carbs) if carbs else 0,
                'protein': float(protein) if protein else 0,
                'fat': float(fat) if fat else 0
            })
    
    return nutritional_data

# Convert sample recipe
def convert_sample_recipe():
    ws = wb['RECETA 1']
    
    # Get recipe name from row 2
    recipe_name = str(ws.cell(2, 2).value or 'Receta 1')
    
    ingredients = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1]:  # No ingredient name
            continue
        
        name = str(row[1]).strip()
        if name == 'TOTALES':
            break
            
        net = row[2]
        unit = row[3]
        cf = row[4]
        gross = row[5]
        cost = row[6]
        
        if name and net:
            try:
                net_val = float(net) if isinstance(net, (int, float)) else 0
                cf_val = float(cf) if cf and isinstance(cf, (int, float)) else 1.0
                gross_val = float(gross) if gross and isinstance(gross, (int, float)) else 0
                cost_val = float(cost) if cost and isinstance(cost, (int, float)) else 0
                
                ingredients.append({
                    'name': name,
                    'netQuantity': net_val,
                    'unit': str(unit).strip() if unit else 'grs',
                    'correctionFactor': cf_val,
                    'grossQuantity': gross_val,
                    'cost': cost_val
                })
            except (ValueError, TypeError):
                continue
    
    
    return {
        'id': 'recipe_1',
        'name': recipe_name,
        'ingredients': ingredients,
        'portions': 4
    }

# Main conversion
print("Converting Excel data to JSON...")

data = {
    'ingredients': convert_ingredients(),
    'correctionFactors': convert_correction_factors(),
    'nutritionalInfo': convert_nutritional_info(),
    'sampleRecipe': convert_sample_recipe()
}

# Save to JSON files
output_dir = Path('src/data')
output_dir.mkdir(exist_ok=True)

with open(output_dir / 'ingredients.json', 'w', encoding='utf-8') as f:
    json.dump(data['ingredients'], f, indent=2, ensure_ascii=False)

with open(output_dir / 'correctionFactors.json', 'w', encoding='utf-8') as f:
    json.dump(data['correctionFactors'], f, indent=2, ensure_ascii=False)

with open(output_dir / 'nutritionalInfo.json', 'w', encoding='utf-8') as f:
    json.dump(data['nutritionalInfo'], f, indent=2, ensure_ascii=False)

with open(output_dir / 'sampleRecipes.json', 'w', encoding='utf-8') as f:
    json.dump([data['sampleRecipe']], f, indent=2, ensure_ascii=False)

print(f"✓ Converted {len(data['ingredients'])} ingredients")
print(f"✓ Converted {len(data['correctionFactors'])} correction factors")
print(f"✓ Converted {len(data['nutritionalInfo'])} nutritional items")
print(f"✓ Converted 1 sample recipe")
print("\nAll data converted successfully!")
